#!/usr/bin/env python3
"""
Reimport für frühere Monate aus Excel-Trackern.
- Morawitz: Januar, Februar, März, Mai 2026 (NK, BK, VL)
- fach.digital: April, Mai 2026 (NK, BK, VL)
Erzeugt Migration 040.
"""
import openpyxl, sqlite3, os, datetime, re

FACH = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/Angebots-Tracker_fach.digital_2026 (4).xlsx'
MOR  = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/Angebots-Tracker Morawitz Consulting NK Sales _ Key Account (1).xlsx'
DB   = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/kpi-tracker/backend/data/kpi.db'
OUT  = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/kpi-tracker/backend/migrations/040_earlier_months_reimport.sql'

EMP_BN = {
    'tobias': 9, 'julius': 12, 'mikail': 10, 'clemens': 11,
    'brian': 15, 'andreas': 14, 'keanu': 23, 'lyes': 16,
    'christian': 29, 'dominik': 32, 'albert': 5, 'serkan': 13,
}
EMP_BS = {
    'bruno': 25, 'marcel': 24, 'tim': 27, 'jerry': 28,
    'keanu': 23, 'christian': 29, 'thorsten': 34,
}
EMP_FACH_NK = {**EMP_BN, **EMP_BS}
EMP_BK_FACH = {'rene': 4, 'joshua': 7, 'daniel': 6, 'sergio': 8, 'adrian': 33, 'albert': 5}
EMP_VL_FACH = {'rene': 4, 'joshua': 7, 'daniel': 6, 'sergio': 8, 'adrian': 33, 'albert': 5}
EMP_MOR_NK = {
    'stefan morawitz': 35, 'sükrü boydas': 36, 'manuel zimmermann': 37,
    'sukrü boydas': 36, 'sukru boydas': 36,
}
EMP_MOR_BK = {
    'stefan morawitz': 35, 'lukas riegler': 39,
    'matthias niedermoser': 41, 'alexander korak': 43,
    'arton mustafa': 40, 'dominik warscha': 38,
}
EMP_MOR_VL = {'alexander korak': 43, 'lukas riegler': 39, 'matthias niedermoser': 41, 'stefan morawitz': 35}

def lookup(name, mapping, fallback=None):
    if not name: return fallback
    k = str(name).strip().lower()
    if k in mapping: return mapping[k]
    for key, val in mapping.items():
        if k in key or key in k: return val
    return fallback

def to_status(v):
    if not v: return 'Offen'
    s = str(v).strip().lower()
    if s == 'ja': return 'Gewonnen'
    if s in ('nein', 'kündigung', 'kuendigung'): return 'Verloren'
    return 'Offen'

def to_date(val, monat):
    if not val: return f'{monat}-01'
    if isinstance(val, datetime.datetime): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    m = re.match(r'(\d{4}-\d{2}-\d{2})', s)
    if m: return m.group(1)
    return f'{monat}-01'

def is_valid_date(val):
    if isinstance(val, (datetime.datetime, datetime.date)): return True
    if not val: return False
    s = str(val).strip()
    return bool(re.match(r'\d{4}-\d{2}-\d{2}', s) or re.match(r'\d{1,2}[./]\d{1,2}[./]\d{2,4}', s))

def to_float(val):
    if val is None: return None
    try: return float(val)
    except: return None

def to_laufzeit(val):
    if val is None: return None
    s = str(val).strip().lower()
    m = re.match(r'(\d+)\s*wochen?', s)
    if m: return round(int(m.group(1)) / 4, 1)
    try: return float(val)
    except: return None

def sql_val(v):
    if v is None: return 'NULL'
    if isinstance(v, float): return repr(v)
    if isinstance(v, int): return str(v)
    return "'" + str(v).replace("'", "''") + "'"

def read_sheet(ws, data_start, col_map, datum_col=None):
    kunde_col = col_map.get('kunde', 3)
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        k_idx = kunde_col - 1
        kunde = row[k_idx] if len(row) > k_idx else None
        if not kunde:
            continue
        if datum_col is not None:
            d_idx = datum_col - 1
            datum_raw = row[d_idx] if len(row) > d_idx else None
            if not is_valid_date(datum_raw):
                continue
        d = {}
        for field, col in col_map.items():
            d[field] = row[col - 1] if len(row) >= col else None
        rows.append(d)
    return rows

# ── SQL-Generator-Funktionen ───────────────────────────────────────────────────

def gen_nk_fach_bn(rows, monat):
    lines = []
    for r in rows:
        status = to_status(r.get('realisierung'))
        wert = to_float(r.get('wert'))
        ae_wert = wert if status == 'Gewonnen' else 0.0
        datum = to_date(r.get('datum'), monat)
        closer_id = lookup(r.get('closer'), EMP_BN) or lookup(r.get('closer'), EMP_BS)
        opener_id = lookup(r.get('opener'), EMP_FACH_NK)
        setter_id = lookup(r.get('setter'), EMP_FACH_NK)
        quelle = str(r.get('quelle') or '').strip() or None
        laufzeit = to_laufzeit(r.get('laufzeit'))
        cols = 'datum,monat,company_id,closer_id,opener_id,setter_id,quelle,kunde,angebotsnummer,dienstleistung,angebotswert,laufzeit_monate,status,ae_wert,kommentar'
        vals = ','.join([
            sql_val(datum), sql_val(monat), '1',
            sql_val(closer_id), sql_val(opener_id), sql_val(setter_id),
            sql_val(quelle), sql_val(r.get('kunde')),
            sql_val(r.get('angebotsnummer')), sql_val(r.get('dienstleistung')),
            sql_val(wert), sql_val(laufzeit),
            sql_val(status), sql_val(ae_wert), sql_val(r.get('kommentar')),
        ])
        lines.append(f'INSERT INTO deals_nk ({cols}) VALUES ({vals});')
    return lines

def gen_nk_fach_bs(rows, monat):
    lines = []
    for r in rows:
        status = to_status(r.get('realisierung'))
        wert = to_float(r.get('wert'))
        ae_wert = wert if status == 'Gewonnen' else 0.0
        datum = to_date(r.get('datum'), monat)
        closer_id = lookup(r.get('closer'), EMP_BS) or lookup(r.get('closer'), EMP_BN)
        opener_id = lookup(r.get('opener'), EMP_FACH_NK)
        setter_id = lookup(r.get('setter'), EMP_FACH_NK)
        quelle = str(r.get('quelle') or '').strip() or None
        laufzeit = to_laufzeit(r.get('laufzeit'))
        cols = 'datum,monat,company_id,closer_id,opener_id,setter_id,quelle,kunde,dienstleistung,angebotswert,laufzeit_monate,status,ae_wert,kommentar'
        vals = ','.join([
            sql_val(datum), sql_val(monat), '1',
            sql_val(closer_id), sql_val(opener_id), sql_val(setter_id),
            sql_val(quelle), sql_val(r.get('kunde')),
            sql_val(r.get('dienstleistung')), sql_val(wert),
            sql_val(laufzeit), sql_val(status), sql_val(ae_wert),
            sql_val(r.get('kommentar')),
        ])
        lines.append(f'INSERT INTO deals_nk ({cols}) VALUES ({vals});')
    return lines

def gen_nk_mor(rows, monat):
    lines = []
    for r in rows:
        status = to_status(r.get('realisierung'))
        wert = to_float(r.get('wert'))
        ae_wert = wert if status == 'Gewonnen' else 0.0
        datum = to_date(r.get('datum'), monat)
        closer_id = lookup(r.get('closer'), EMP_MOR_NK)
        quelle = str(r.get('quelle') or '').strip() or None
        laufzeit = to_laufzeit(r.get('laufzeit'))
        cols = 'datum,monat,company_id,closer_id,quelle,kunde,angebotsnummer,dienstleistung,angebotswert,laufzeit_monate,status,ae_wert,kommentar'
        vals = ','.join([
            sql_val(datum), sql_val(monat), '3',
            sql_val(closer_id), sql_val(quelle),
            sql_val(r.get('kunde')), sql_val(r.get('angebotsnummer')),
            sql_val(r.get('dienstleistung')), sql_val(wert),
            sql_val(laufzeit), sql_val(status), sql_val(ae_wert),
            sql_val(r.get('kommentar')),
        ])
        lines.append(f'INSERT INTO deals_nk ({cols}) VALUES ({vals});')
    return lines

def gen_bk_fach(rows, monat):
    lines = []
    for r in rows:
        status = to_status(r.get('realisierung'))
        wert = to_float(r.get('wert'))
        ae_wert = wert if status == 'Gewonnen' else 0.0
        datum = to_date(r.get('datum'), monat)
        kam_id = lookup(r.get('kam'), EMP_BK_FACH)
        laufzeit = to_laufzeit(r.get('laufzeit'))
        cols = 'datum,monat,company_id,kam_id,kunde,angebotsnummer,dienstleistung,angebotswert,laufzeit_monate,status,ae_wert,kommentar'
        vals = ','.join([
            sql_val(datum), sql_val(monat), '1',
            sql_val(kam_id), sql_val(r.get('kunde')),
            sql_val(r.get('angebotsnummer')), sql_val(r.get('dienstleistung')),
            sql_val(wert), sql_val(laufzeit),
            sql_val(status), sql_val(ae_wert), sql_val(r.get('kommentar')),
        ])
        lines.append(f'INSERT INTO deals_bk ({cols}) VALUES ({vals});')
    return lines

def gen_bk_mor(rows, monat):
    lines = []
    for r in rows:
        status = to_status(r.get('realisierung'))
        wert = to_float(r.get('wert'))
        ae_wert = wert if status == 'Gewonnen' else 0.0
        datum = to_date(r.get('datum'), monat)
        kam_id = lookup(r.get('kam'), EMP_MOR_BK)
        laufzeit = to_laufzeit(r.get('laufzeit'))
        cols = 'datum,monat,company_id,kam_id,kunde,angebotsnummer,dienstleistung,angebotswert,laufzeit_monate,status,ae_wert,kommentar'
        vals = ','.join([
            sql_val(datum), sql_val(monat), '3',
            sql_val(kam_id), sql_val(r.get('kunde')),
            sql_val(r.get('angebotsnummer')), sql_val(r.get('dienstleistung')),
            sql_val(wert), sql_val(laufzeit),
            sql_val(status), sql_val(ae_wert), sql_val(r.get('kommentar')),
        ])
        lines.append(f'INSERT INTO deals_bk ({cols}) VALUES ({vals});')
    return lines

def gen_vl(rows, monat, company_id, emp_map):
    lines = []
    for r in rows:
        status = to_status(r.get('realisierung'))
        ae_raw = to_float(r.get('ae'))
        kont = to_float(r.get('kontingentpreis'))
        ae_wert = ae_raw if status == 'Gewonnen' else 0.0
        angebotswert = ae_raw if ae_raw else kont
        datum = to_date(r.get('datum'), monat)
        kam_id = lookup(r.get('am'), emp_map)
        laufzeit = to_laufzeit(r.get('laufzeit'))
        wie_vielt = to_float(r.get('wie_vielt'))
        cols = 'datum,monat,company_id,kam_id,kunde,dienstleistung,angebotswert,ae_wert,laufzeit_monate,status,wie_vielt_verlaengerung'
        vals = ','.join([
            sql_val(datum), sql_val(monat), str(company_id),
            sql_val(kam_id), sql_val(r.get('kunde')),
            sql_val(r.get('dienstleistung')), sql_val(angebotswert),
            sql_val(ae_wert), sql_val(laufzeit),
            sql_val(status), sql_val(wie_vielt),
        ])
        lines.append(f'INSERT INTO deals_vl ({cols}) VALUES ({vals});')
    return lines

# ── Excel öffnen ──────────────────────────────────────────────────────────────
print('Öffne Excel-Dateien...')
wb_f = openpyxl.load_workbook(FACH, data_only=True)
wb_m = openpyxl.load_workbook(MOR, data_only=True)

# ── Morawitz NK lesen ─────────────────────────────────────────────────────────
# Jan-März: Realisierung=Col8 (kein Quelle-Feld)
# Mai: Quelle=Col8, Realisierung=Col9
NK_MOR_JAN_MAR = {'datum':1,'closer':2,'kunde':3,'angebotsnummer':4,'dienstleistung':5,'wert':6,'laufzeit':7,'realisierung':8}
NK_MOR_MAI     = {'datum':1,'closer':2,'kunde':3,'angebotsnummer':4,'dienstleistung':5,'wert':6,'laufzeit':7,'quelle':8,'realisierung':9}

mor_nk = {
    '2026-01': read_sheet(wb_m['NK Januar 26'],   13, NK_MOR_JAN_MAR),
    '2026-02': read_sheet(wb_m['NK Februar 26'],  13, NK_MOR_JAN_MAR),
    '2026-03': read_sheet(wb_m['NK März 26'],     13, NK_MOR_JAN_MAR),
    '2026-05': read_sheet(wb_m['NK Mai 26'],      13, NK_MOR_MAI),
}
for m, rows in mor_nk.items(): print(f'  NK Morawitz {m}: {len(rows)} Zeilen')

# ── Morawitz BK lesen ─────────────────────────────────────────────────────────
# Jan-März: Datum=1, ..., Realisierung=8, Kommentar=9
# Mai: +Besprechungstermin bei Col8 → Realisierung=9, Kommentar=10
BK_MOR_JAN_MAR = {'datum':1,'kam':2,'kunde':3,'angebotsnummer':4,'dienstleistung':5,'wert':6,'laufzeit':7,'realisierung':8,'kommentar':9}
BK_MOR_MAI     = {'datum':1,'kam':2,'kunde':3,'angebotsnummer':4,'dienstleistung':5,'wert':6,'laufzeit':7,'realisierung':9,'kommentar':10}

mor_bk = {
    '2026-01': read_sheet(wb_m['Key Acc. Januar 26'],   13, BK_MOR_JAN_MAR, datum_col=1),
    '2026-02': read_sheet(wb_m['Key Acc. Februar 26'],  13, BK_MOR_JAN_MAR, datum_col=1),
    '2026-03': read_sheet(wb_m['Key Acc. März 26'],     13, BK_MOR_JAN_MAR, datum_col=1),
    '2026-05': read_sheet(wb_m['Key Acc. Mai 26'],      13, BK_MOR_MAI,     datum_col=1),
}
for m, rows in mor_bk.items(): print(f'  BK Morawitz {m}: {len(rows)} Zeilen')

# ── Morawitz VL lesen ─────────────────────────────────────────────────────────
# Alle Monate: konsistent, Header R12, Daten ab R13
VL_MOR = {'datum':1,'kunde':2,'am':3,'dienstleistung':4,'laufzeit':6,'kontingentpreis':7,'ae':8,'realisierung':9}

mor_vl = {
    '2026-01': read_sheet(wb_m['Auto. Verlängerungen Jan 26'],  13, VL_MOR),
    '2026-02': read_sheet(wb_m['Auto. Verlängerungen Febr 26'], 13, VL_MOR),
    '2026-03': read_sheet(wb_m['Auto. Verlängerungen März 26'], 13, VL_MOR),
    '2026-05': read_sheet(wb_m['Auto. Verlängerungen Mai 26'],  13, VL_MOR),
}
for m, rows in mor_vl.items(): print(f'  VL Morawitz {m}: {len(rows)} Zeilen')

# ── fach.digital NK April ─────────────────────────────────────────────────────
# Ein kombiniertes Sheet (kein BN/BS-Split), BN-Struktur (Realisierung=13)
# Header R17, Daten ab R18
NK_FACH_BN = {
    'quelle':1,'datum':2,'closer':3,'opener':4,'setter':5,
    'kunde':6,'angebotsnummer':8,'dienstleistung':9,
    'wert':10,'laufzeit':11,'kommentar':12,'realisierung':13,
}
NK_FACH_BS = {
    'quelle':1,'datum':2,'closer':3,'opener':4,'setter':5,
    'kunde':6,'dienstleistung':8,'wert':9,
    'laufzeit':10,'kommentar':11,'realisierung':12,
}

fach_nk_apr    = read_sheet(wb_f['NK-Vertrieb April 2026'],   18, NK_FACH_BN)
fach_nk_mai_bn = read_sheet(wb_f['NK-Vertrieb Mai BN 2026'], 18, NK_FACH_BN)
fach_nk_mai_bs = read_sheet(wb_f['NK-Vertrieb Mai BS 2026'], 18, NK_FACH_BS)
print(f'  NK fach.digital April: {len(fach_nk_apr)}, Mai BN: {len(fach_nk_mai_bn)}, Mai BS: {len(fach_nk_mai_bs)}')

# ── fach.digital BK April+Mai ─────────────────────────────────────────────────
# Header R14, Daten ab R15, Realisierung=9 (beide Monate)
BK_FACH = {'datum':1,'kam':2,'kunde':3,'angebotsnummer':4,'dienstleistung':5,'wert':6,'laufzeit':7,'kommentar':8,'realisierung':9}

fach_bk_apr = read_sheet(wb_f['BK April 2026'], 15, BK_FACH)
fach_bk_mai = read_sheet(wb_f['BK Mai 2026'],   15, BK_FACH)
print(f'  BK fach.digital April: {len(fach_bk_apr)}, Mai: {len(fach_bk_mai)}')

# ── fach.digital VL April+Mai ─────────────────────────────────────────────────
# Header R12, Daten ab R13 (verschoben vs. Juni R11/R12!)
VL_FACH = {'datum':1,'kunde':2,'am':3,'dienstleistung':4,'laufzeit':6,'kontingentpreis':7,'ae':8,'realisierung':9,'wie_vielt':10}

fach_vl_apr = read_sheet(wb_f['Auto. Verlängerungen April 2026'], 13, VL_FACH)
fach_vl_mai = read_sheet(wb_f['Auto. Verlängerungen Mai 2026'],   13, VL_FACH)
print(f'  VL fach.digital April: {len(fach_vl_apr)}, Mai: {len(fach_vl_mai)}')

# ── SQL generieren ─────────────────────────────────────────────────────────────
lines = []
lines.append('-- Migration 040: Reimport früherer Monate aus Excel-Trackern')
lines.append('-- Morawitz: Januar, Februar, März, Mai 2026 (NK, BK, VL)')
lines.append('-- fach.digital: April, Mai 2026 (NK, BK, VL)')
lines.append('')

# DELETEs – nur betroffene Firmen/Monate
lines.append('-- DELETEs: nur betroffene company_id/Monate')
for m in ('2026-01', '2026-02', '2026-03', '2026-05'):
    lines.append(f"DELETE FROM deals_nk WHERE monat='{m}' AND company_id=3;")
    lines.append(f"DELETE FROM deals_bk WHERE monat='{m}' AND company_id=3;")
    lines.append(f"DELETE FROM deals_vl WHERE monat='{m}' AND company_id=3;")
for m in ('2026-04', '2026-05'):
    lines.append(f"DELETE FROM deals_nk WHERE monat='{m}' AND company_id=1;")
    lines.append(f"DELETE FROM deals_bk WHERE monat='{m}' AND company_id=1;")
    lines.append(f"DELETE FROM deals_vl WHERE monat='{m}' AND company_id=1;")
lines.append('')

# Morawitz NK
for monat, rows in mor_nk.items():
    lines.append(f'-- NK Morawitz {monat} ({len(rows)} Zeilen)')
    lines.extend(gen_nk_mor(rows, monat))
    lines.append('')

# fach.digital NK April (BN-Struktur, ein Sheet)
lines.append(f'-- NK fach.digital April 2026 ({len(fach_nk_apr)} Zeilen, BN-Struktur)')
lines.extend(gen_nk_fach_bn(fach_nk_apr, '2026-04'))
lines.append('')

# fach.digital NK Mai (getrennt BN/BS)
lines.append(f'-- NK fach.digital Mai 2026 BN ({len(fach_nk_mai_bn)} Zeilen)')
lines.extend(gen_nk_fach_bn(fach_nk_mai_bn, '2026-05'))
lines.append('')
lines.append(f'-- NK fach.digital Mai 2026 BS ({len(fach_nk_mai_bs)} Zeilen)')
lines.extend(gen_nk_fach_bs(fach_nk_mai_bs, '2026-05'))
lines.append('')

# Morawitz BK
for monat, rows in mor_bk.items():
    lines.append(f'-- BK Morawitz {monat} ({len(rows)} Zeilen)')
    lines.extend(gen_bk_mor(rows, monat))
    lines.append('')

# fach.digital BK
lines.append(f'-- BK fach.digital April 2026 ({len(fach_bk_apr)} Zeilen)')
lines.extend(gen_bk_fach(fach_bk_apr, '2026-04'))
lines.append('')
lines.append(f'-- BK fach.digital Mai 2026 ({len(fach_bk_mai)} Zeilen)')
lines.extend(gen_bk_fach(fach_bk_mai, '2026-05'))
lines.append('')

# Morawitz VL
for monat, rows in mor_vl.items():
    lines.append(f'-- VL Morawitz {monat} ({len(rows)} Zeilen)')
    lines.extend(gen_vl(rows, monat, 3, EMP_MOR_VL))
    lines.append('')

# fach.digital VL
lines.append(f'-- VL fach.digital April 2026 ({len(fach_vl_apr)} Zeilen)')
lines.extend(gen_vl(fach_vl_apr, '2026-04', 1, EMP_VL_FACH))
lines.append('')
lines.append(f'-- VL fach.digital Mai 2026 ({len(fach_vl_mai)} Zeilen)')
lines.extend(gen_vl(fach_vl_mai, '2026-05', 1, EMP_VL_FACH))
lines.append('')

# gewonnen_monat setzen für Live-Standort-Abfragen
lines.append('-- gewonnen_monat setzen (alle betroffenen Monate + Firmen)')
affected = ("'2026-01','2026-02','2026-03','2026-04','2026-05'")
lines.append(f"UPDATE deals_nk SET gewonnen_monat=monat, gewonnen_datum=datum WHERE monat IN ({affected}) AND status='Gewonnen' AND company_id IN (1,3) AND (gewonnen_monat IS NULL OR gewonnen_monat='');")
lines.append(f"UPDATE deals_bk SET gewonnen_monat=monat, gewonnen_datum=datum WHERE monat IN ({affected}) AND status='Gewonnen' AND company_id IN (1,3) AND (gewonnen_monat IS NULL OR gewonnen_monat='');")
lines.append(f"UPDATE deals_vl SET gewonnen_monat=monat, gewonnen_datum=datum WHERE monat IN ({affected}) AND status='Gewonnen' AND company_id IN (1,3) AND (gewonnen_monat IS NULL OR gewonnen_monat='');")
lines.append('')

# ── Lokal anwenden ────────────────────────────────────────────────────────────
deal_sql = '\n'.join(lines) + '\n'

print('\n=== Wende Deal-SQL auf lokale DB an... ===')
conn = sqlite3.connect(DB)
conn.executescript('PRAGMA foreign_keys=OFF;\n' + deal_sql)
conn.commit()

# ── ae_gesamt_monthly aktualisieren ───────────────────────────────────────────
# BK/VL Gesamtsummen aus DB abfragen und ae_gesamt_monthly anpassen
# NK-Werte bleiben UNVERÄNDERT (AE Gesamt ist autoritär)
ag_lines = []
ag_lines.append('-- ae_gesamt_monthly: BK/VL-Totale aktualisieren basierend auf neuen Deal-Daten')

affected_months = ['2026-01', '2026-02', '2026-03', '2026-04', '2026-05']
for monat in affected_months:
    bk_total = conn.execute(
        "SELECT COALESCE(SUM(ae_wert),0) FROM deals_bk WHERE monat=? AND status='Gewonnen'", (monat,)
    ).fetchone()[0]
    vl_total = conn.execute(
        "SELECT COALESCE(SUM(ae_wert),0) FROM deals_vl WHERE monat=? AND status='Gewonnen'", (monat,)
    ).fetchone()[0]
    row = conn.execute(
        "SELECT nk_gesamt, gesamt FROM ae_gesamt_monthly WHERE monat=?", (monat,)
    ).fetchone()
    if not row:
        print(f'  WARNUNG: Kein ae_gesamt_monthly Eintrag für {monat}')
        continue
    nk_gesamt = row[0] or 0
    new_gesamt = round(nk_gesamt + bk_total + vl_total, 2)
    ag_lines.append(f"UPDATE ae_gesamt_monthly SET bk_gesamt={bk_total}, vl_gesamt={vl_total}, gesamt={new_gesamt} WHERE monat='{monat}';")
    print(f'  {monat}: NK={nk_gesamt:,.0f} + BK={bk_total:,.0f} + VL={vl_total:,.0f} = Gesamt={new_gesamt:,.0f}')

# ae_gesamt_monthly Updates anwenden
conn.executescript('\n'.join(ag_lines))
conn.commit()

# ── Migration-Datei schreiben ─────────────────────────────────────────────────
full_sql = deal_sql + '\n'.join(ag_lines) + '\n'
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(full_sql)

# ── Verifikation ──────────────────────────────────────────────────────────────
print('\n=== Verifikation nach Import ===')
for monat in affected_months:
    print(f'\n  {monat}:')
    for tbl, standort_q, cid_label in [
        ('deals_nk', "JOIN employees e ON deals_nk.closer_id=e.id", None),
        ('deals_bk', None, None),
        ('deals_vl', None, None),
    ]:
        total = conn.execute(f"SELECT COUNT(*), COALESCE(SUM(ae_wert),0) FROM {tbl} WHERE monat=? AND status='Gewonnen' AND company_id IN (1,3)", (monat,)).fetchone()
        print(f'    {tbl}: {total[0]} Gewonnen, {total[1]:,.0f}€ AE')

conn.close()

total_inserts = (
    sum(len(r) for r in mor_nk.values()) +
    sum(len(r) for r in mor_bk.values()) +
    sum(len(r) for r in mor_vl.values()) +
    len(fach_nk_apr) + len(fach_nk_mai_bn) + len(fach_nk_mai_bs) +
    len(fach_bk_apr) + len(fach_bk_mai) +
    len(fach_vl_apr) + len(fach_vl_mai)
)
print(f'\n=== Gesamt: {total_inserts} INSERTs ===')
print(f'Geschrieben: {OUT}')
