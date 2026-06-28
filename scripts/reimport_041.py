#!/usr/bin/env python3
"""
Migration 041:
- Morawitz April 2026: NK, BK, VL vollständig reimportieren
- fach.digital NK Februar + März: Reimport (Verloren-Status war falsch)
"""
import openpyxl, sqlite3, os, datetime, re

FACH = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/Angebots-Tracker_fach.digital_2026 (4).xlsx'
MOR  = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/Angebots-Tracker Morawitz Consulting NK Sales _ Key Account (1).xlsx'
DB   = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/kpi-tracker/backend/data/kpi.db'
OUT  = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/kpi-tracker/backend/migrations/041_april_febmar_fix.sql'

EMP_BN = {
    'tobias': 9, 'julius': 12, 'mikail': 10, 'clemens': 11,
    'brian': 15, 'andreas': 14, 'keanu': 23, 'lyes': 16,
    'christian': 29, 'dominik': 32, 'albert': 5, 'serkan': 13,
}
EMP_BS = {
    'bruno': 25, 'marcel': 24, 'tim': 27, 'jerry': 28,
    'keanu': 23, 'christian': 29, 'thorsten': 34,
}
EMP_FACH_NK = {**EMP_BN, **EMP_BS}
EMP_BK_FACH = {'rene': 4, 'joshua': 7, 'daniel': 6, 'sergio': 8, 'adrian': 33, 'albert': 5}
EMP_VL_FACH = {'rene': 4, 'joshua': 7, 'daniel': 6, 'sergio': 8, 'adrian': 33, 'albert': 5}
EMP_MOR_NK  = {
    'stefan morawitz': 35, 'sükrü boydas': 36, 'manuel zimmermann': 37,
    'sukrü boydas': 36, 'sukru boydas': 36,
}
EMP_MOR_BK  = {
    'stefan morawitz': 35, 'lukas riegler': 39,
    'matthias niedermoser': 41, 'alexander korak': 43,
    'arton mustafa': 40, 'dominik warscha': 38,
    'daniel saliba': 42,
}
EMP_MOR_VL  = {'alexander korak': 43, 'lukas riegler': 39, 'matthias niedermoser': 41, 'stefan morawitz': 35}

def lookup(name, mapping, fallback=None):
    if not name: return fallback
    k = str(name).strip().lower()
    if k in mapping: return mapping[k]
    for key, val in mapping.items():
        if k in key or key in k: return val
    return fallback

def to_status(v):
    if not v: return 'Offen'
    s = str(v).strip().lower()
    if s == 'ja': return 'Gewonnen'
    if s in ('nein', 'kündigung', 'kuendigung'): return 'Verloren'
    return 'Offen'

def to_date(val, monat):
    if not val: return f'{monat}-01'
    if isinstance(val, datetime.datetime): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    m = re.match(r'(\d{4}-\d{2}-\d{2})', s)
    if m: return m.group(1)
    return f'{monat}-01'

def is_valid_date(val):
    if isinstance(val, (datetime.datetime, datetime.date)): return True
    if not val: return False
    s = str(val).strip()
    return bool(re.match(r'\d{4}-\d{2}-\d{2}', s) or re.match(r'\d{1,2}[./]\d{1,2}[./]\d{2,4}', s))

def to_float(val):
    if val is None: return None
    try: return float(val)
    except: return None

def to_laufzeit(val):
    if val is None: return None
    s = str(val).strip().lower()
    m = re.match(r'(\d+)\s*wochen?', s)
    if m: return round(int(m.group(1)) / 4, 1)
    try: return float(val)
    except: return None

def sql_val(v):
    if v is None: return 'NULL'
    if isinstance(v, float): return repr(v)
    if isinstance(v, int): return str(v)
    return "'" + str(v).replace("'", "''") + "'"

def read_sheet(ws, data_start, col_map, datum_col=None):
    kunde_col = col_map.get('kunde', 3)
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        # Stopp bei Provision-Tabelle (nur Col1 prüfen)
        if row and row[0] and 'provision' in str(row[0]).lower():
            break
        k_idx = kunde_col - 1
        kunde = row[k_idx] if len(row) > k_idx else None
        if not kunde:
            continue
        if datum_col is not None:
            d_idx = datum_col - 1
            datum_raw = row[d_idx] if len(row) > d_idx else None
            if not is_valid_date(datum_raw):
                continue
        d = {}
        for field, col in col_map.items():
            d[field] = row[col - 1] if len(row) >= col else None
        rows.append(d)
    return rows

# ── Spaltenmappings ───────────────────────────────────────────────────────────
# Morawitz NK/BK April: wie Jan-März (Realisierung=Col8)
NK_MOR_JAN_MAR = {'datum':1,'closer':2,'kunde':3,'angebotsnummer':4,'dienstleistung':5,'wert':6,'laufzeit':7,'realisierung':8}
BK_MOR_JAN_MAR = {'datum':1,'kam':2,'kunde':3,'angebotsnummer':4,'dienstleistung':5,'wert':6,'laufzeit':7,'realisierung':8,'kommentar':9}
VL_MOR         = {'datum':1,'kunde':2,'am':3,'dienstleistung':4,'laufzeit':6,'kontingentpreis':7,'ae':8,'realisierung':9}
# fach.digital NK (alle kombinierten Sheets Jan-April): Header R17, Data R18, Reak=Col13
NK_FACH_BN = {
    'quelle':1,'datum':2,'closer':3,'opener':4,'setter':5,
    'kunde':6,'angebotsnummer':8,'dienstleistung':9,
    'wert':10,'laufzeit':11,'kommentar':12,'realisierung':13,
}

print('Öffne Excel-Dateien...')
wb_f = openpyxl.load_workbook(FACH, data_only=True)
wb_m = openpyxl.load_workbook(MOR, data_only=True)

# ── Morawitz April lesen ──────────────────────────────────────────────────────
print('Lese Morawitz April...')
mor_apr_nk = read_sheet(wb_m['NK April 26'],                    13, NK_MOR_JAN_MAR)
mor_apr_bk = read_sheet(wb_m['Key Acc. April 26'],              13, BK_MOR_JAN_MAR, datum_col=None)
mor_apr_vl = read_sheet(wb_m['Auto. Verlängerungen April 26'],  13, VL_MOR)
print(f'  NK: {len(mor_apr_nk)}, BK: {len(mor_apr_bk)}, VL: {len(mor_apr_vl)}')

# ── fach.digital NK Februar + März lesen ─────────────────────────────────────
print('Lese fach.digital NK Feb + März...')
fach_nk_feb = read_sheet(wb_f['NK-Vertrieb Februar 26'], 18, NK_FACH_BN)
fach_nk_mar = read_sheet(wb_f['NK-Vertrieb März 26'],    18, NK_FACH_BN)
print(f'  Feb: {len(fach_nk_feb)}, März: {len(fach_nk_mar)}')

# ── SQL generieren ─────────────────────────────────────────────────────────────
lines = []
lines.append('-- Migration 041: Morawitz April + fach.digital NK Feb/März Korrektur')
lines.append('')

# DELETEs
lines.append('-- DELETEs')
lines.append("DELETE FROM deals_nk WHERE monat='2026-04' AND company_id=3;")
lines.append("DELETE FROM deals_bk WHERE monat='2026-04' AND company_id=3;")
lines.append("DELETE FROM deals_vl WHERE monat='2026-04' AND company_id=3;")
lines.append("DELETE FROM deals_nk WHERE monat='2026-02' AND company_id=1;")
lines.append("DELETE FROM deals_nk WHERE monat='2026-03' AND company_id=1;")
lines.append('')

# Morawitz April NK
MONAT_APR = '2026-04'
lines.append(f'-- NK Morawitz April ({len(mor_apr_nk)} Zeilen)')
for r in mor_apr_nk:
    status = to_status(r.get('realisierung'))
    wert = to_float(r.get('wert'))
    ae_wert = wert if status == 'Gewonnen' else 0.0
    datum = to_date(r.get('datum'), MONAT_APR)
    closer_id = lookup(r.get('closer'), EMP_MOR_NK)
    laufzeit = to_laufzeit(r.get('laufzeit'))
    cols = 'datum,monat,company_id,closer_id,quelle,kunde,angebotsnummer,dienstleistung,angebotswert,laufzeit_monate,status,ae_wert,kommentar'
    vals = ','.join([
        sql_val(datum), sql_val(MONAT_APR), '3',
        sql_val(closer_id), 'NULL',
        sql_val(r.get('kunde')), sql_val(r.get('angebotsnummer')),
        sql_val(r.get('dienstleistung')), sql_val(wert),
        sql_val(laufzeit), sql_val(status), sql_val(ae_wert), 'NULL',
    ])
    lines.append(f'INSERT INTO deals_nk ({cols}) VALUES ({vals});')
lines.append('')

# Morawitz April BK
lines.append(f'-- BK Morawitz April ({len(mor_apr_bk)} Zeilen)')
for r in mor_apr_bk:
    status = to_status(r.get('realisierung'))
    wert = to_float(r.get('wert'))
    ae_wert = wert if status == 'Gewonnen' else 0.0
    datum = to_date(r.get('datum'), MONAT_APR)
    kam_id = lookup(r.get('kam'), EMP_MOR_BK)
    laufzeit = to_laufzeit(r.get('laufzeit'))
    cols = 'datum,monat,company_id,kam_id,kunde,angebotsnummer,dienstleistung,angebotswert,laufzeit_monate,status,ae_wert,kommentar'
    vals = ','.join([
        sql_val(datum), sql_val(MONAT_APR), '3',
        sql_val(kam_id), sql_val(r.get('kunde')),
        sql_val(r.get('angebotsnummer')), sql_val(r.get('dienstleistung')),
        sql_val(wert), sql_val(laufzeit),
        sql_val(status), sql_val(ae_wert), sql_val(r.get('kommentar')),
    ])
    lines.append(f'INSERT INTO deals_bk ({cols}) VALUES ({vals});')
lines.append('')

# Morawitz April VL
lines.append(f'-- VL Morawitz April ({len(mor_apr_vl)} Zeilen)')
for r in mor_apr_vl:
    status = to_status(r.get('realisierung'))
    ae_raw = to_float(r.get('ae'))
    kont = to_float(r.get('kontingentpreis'))
    ae_wert = ae_raw if status == 'Gewonnen' else 0.0
    angebotswert = ae_raw if ae_raw else kont
    datum = to_date(r.get('datum'), MONAT_APR)
    kam_id = lookup(r.get('am'), EMP_MOR_VL)
    laufzeit = to_laufzeit(r.get('laufzeit'))
    cols = 'datum,monat,company_id,kam_id,kunde,dienstleistung,angebotswert,ae_wert,laufzeit_monate,status,wie_vielt_verlaengerung'
    vals = ','.join([
        sql_val(datum), sql_val(MONAT_APR), '3',
        sql_val(kam_id), sql_val(r.get('kunde')),
        sql_val(r.get('dienstleistung')), sql_val(angebotswert),
        sql_val(ae_wert), sql_val(laufzeit), sql_val(status), 'NULL',
    ])
    lines.append(f'INSERT INTO deals_vl ({cols}) VALUES ({vals});')
lines.append('')

# fach.digital NK Februar
def gen_fach_nk_combined(rows, monat):
    result = []
    for r in rows:
        status = to_status(r.get('realisierung'))
        wert = to_float(r.get('wert'))
        ae_wert = wert if status == 'Gewonnen' else 0.0
        datum = to_date(r.get('datum'), monat)
        closer_id = lookup(r.get('closer'), EMP_BN) or lookup(r.get('closer'), EMP_BS)
        opener_id = lookup(r.get('opener'), EMP_FACH_NK)
        setter_id  = lookup(r.get('setter'),  EMP_FACH_NK)
        quelle = str(r.get('quelle') or '').strip() or None
        laufzeit = to_laufzeit(r.get('laufzeit'))
        cols = 'datum,monat,company_id,closer_id,opener_id,setter_id,quelle,kunde,angebotsnummer,dienstleistung,angebotswert,laufzeit_monate,status,ae_wert,kommentar'
        vals = ','.join([
            sql_val(datum), sql_val(monat), '1',
            sql_val(closer_id), sql_val(opener_id), sql_val(setter_id),
            sql_val(quelle), sql_val(r.get('kunde')),
            sql_val(r.get('angebotsnummer')), sql_val(r.get('dienstleistung')),
            sql_val(wert), sql_val(laufzeit),
            sql_val(status), sql_val(ae_wert), sql_val(r.get('kommentar')),
        ])
        result.append(f'INSERT INTO deals_nk ({cols}) VALUES ({vals});')
    return result

lines.append(f'-- NK fach.digital Februar ({len(fach_nk_feb)} Zeilen)')
lines.extend(gen_fach_nk_combined(fach_nk_feb, '2026-02'))
lines.append('')

lines.append(f'-- NK fach.digital März ({len(fach_nk_mar)} Zeilen)')
lines.extend(gen_fach_nk_combined(fach_nk_mar, '2026-03'))
lines.append('')

# gewonnen_monat
lines.append('-- gewonnen_monat setzen')
lines.append("UPDATE deals_nk SET gewonnen_monat=monat, gewonnen_datum=datum WHERE monat IN ('2026-02','2026-03','2026-04') AND status='Gewonnen' AND company_id IN (1,3) AND (gewonnen_monat IS NULL OR gewonnen_monat='');")
lines.append("UPDATE deals_bk SET gewonnen_monat=monat, gewonnen_datum=datum WHERE monat='2026-04' AND status='Gewonnen' AND company_id=3 AND (gewonnen_monat IS NULL OR gewonnen_monat='');")
lines.append("UPDATE deals_vl SET gewonnen_monat=monat, gewonnen_datum=datum WHERE monat='2026-04' AND status='Gewonnen' AND company_id=3 AND (gewonnen_monat IS NULL OR gewonnen_monat='');")
lines.append('')

# ── Lokal anwenden ────────────────────────────────────────────────────────────
deal_sql = '\n'.join(lines) + '\n'
print('\n=== Wende SQL auf lokale DB an... ===')
conn = sqlite3.connect(DB)
conn.executescript('PRAGMA foreign_keys=OFF;\n' + deal_sql)
conn.commit()

# ── ae_gesamt_monthly für April aktualisieren ─────────────────────────────────
# Nur BK + VL (NK-Wert bleibt autoritär aus AE Gesamt)
ag_lines = ['-- ae_gesamt_monthly April: BK/VL neu berechnet (NK unverändert)']
bk_total = conn.execute("SELECT COALESCE(SUM(ae_wert),0) FROM deals_bk WHERE monat='2026-04' AND status='Gewonnen'").fetchone()[0]
vl_total = conn.execute("SELECT COALESCE(SUM(ae_wert),0) FROM deals_vl WHERE monat='2026-04' AND status='Gewonnen'").fetchone()[0]
nk_gesamt = conn.execute("SELECT nk_gesamt FROM ae_gesamt_monthly WHERE monat='2026-04'").fetchone()[0] or 0
new_gesamt = round(nk_gesamt + bk_total + vl_total, 2)
ag_lines.append(f"UPDATE ae_gesamt_monthly SET bk_gesamt={bk_total}, vl_gesamt={vl_total}, gesamt={new_gesamt} WHERE monat='2026-04';")
print(f'  April: NK={nk_gesamt:,.0f} + BK={bk_total:,.0f} + VL={vl_total:,.0f} = Gesamt={new_gesamt:,.0f}')

conn.executescript('\n'.join(ag_lines))
conn.commit()

# ── Verifikation ──────────────────────────────────────────────────────────────
print('\n=== Verifikation ===')
for monat, label in [('2026-02','Feb'), ('2026-03','März'), ('2026-04','April')]:
    print(f'\n  {label} fach.digital NK Bonn:')
    rows = conn.execute('''
        SELECT n.status, COUNT(*) FROM deals_nk n
        LEFT JOIN employees e ON n.closer_id=e.id
        WHERE n.monat=? AND n.company_id=1 AND e.standort='Bonn'
        GROUP BY n.status
    ''', (monat,)).fetchall()
    print(f'    {dict(rows)}')

print('\n  April Morawitz:')
for tbl in ('deals_nk', 'deals_bk', 'deals_vl'):
    r = conn.execute(f"SELECT COUNT(*), COUNT(CASE WHEN status='Gewonnen' THEN 1 END), COALESCE(SUM(ae_wert),0) FROM {tbl} WHERE monat='2026-04' AND company_id=3").fetchone()
    print(f'    {tbl}: {r[0]} total, {r[1]} Gewonnen, {r[2]:,.0f}€ AE')
conn.close()

# ── Migration schreiben ───────────────────────────────────────────────────────
full_sql = deal_sql + '\n'.join(ag_lines) + '\n'
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(full_sql)
print(f'\nGeschrieben: {OUT}')
