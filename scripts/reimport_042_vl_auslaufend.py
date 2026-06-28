#!/usr/bin/env python3
"""
Migration 042: VL auslaufend_am + weitergeben_an_vertrieb wiederherstellen.

Alle VL-Sheets, die durch Migrationen 039–041 reimportiert wurden, enthielten
kein auslaufend_am (Col5 = Laufzeit-bis). Diese Migration liest Col5 aus jedem
betroffenen Sheet und setzt auslaufend_am für die entsprechenden Deals.
Außerdem: weitergeben_an_vertrieb='Ja' für alle Verloren-VL-Deals in
betroffenen Monaten setzen (vorheriger Wert war manuell gesetzt und ist verloren).
"""
import openpyxl, sqlite3, os, datetime, re

FACH = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/Angebots-Tracker_fach.digital_2026 (4).xlsx'
MOR  = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/Angebots-Tracker Morawitz Consulting NK Sales _ Key Account (1).xlsx'
DB   = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/kpi-tracker/backend/data/kpi.db'
OUT  = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/kpi-tracker/backend/migrations/042_vl_auslaufend_am.sql'

def to_date_str(val):
    if not val: return None
    if isinstance(val, datetime.datetime): return val.strftime('%Y-%m-%d')
    if isinstance(val, datetime.date): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    m = re.match(r'(\d{4}-\d{2}-\d{2})', s)
    if m: return m.group(1)
    # Try m/d/yyyy
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m: return f'{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
    return None

def sql_val(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def read_vl_auslaufend(ws, data_start, col_kunde, col_auslaufend):
    """Liest Kunde (col_kunde) + Auslaufend-Datum (col_auslaufend) aus VL-Sheet."""
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        k_idx = col_kunde - 1
        a_idx = col_auslaufend - 1
        kunde = row[k_idx] if len(row) > k_idx else None
        if not kunde:
            continue
        auslaufend = row[a_idx] if len(row) > a_idx else None
        rows.append({
            'kunde': str(kunde).strip(),
            'auslaufend_am': to_date_str(auslaufend),
        })
    return rows

print('Öffne Excel-Dateien...')
wb_f = openpyxl.load_workbook(FACH, data_only=True)
wb_m = openpyxl.load_workbook(MOR, data_only=True)

# VL-Sheet Konfigurationen: (wb, sheet_name, monat, company_id, data_start, col_kunde, col_auslaufend)
# Alle VL-Sheets: Kunde=Col2, Laufzeit-bis=Col5
SHEETS = [
    # June (Migration 039)
    (wb_f, 'Auto. Verlängerungen Juni 2026', '2026-06', 1, 12, 2, 5),
    (wb_m, 'Auto. Verlängerungen Juni 26',   '2026-06', 3, 13, 2, 5),
    # May (Migration 040)
    (wb_f, 'Auto. Verlängerungen Mai 2026',  '2026-05', 1, 13, 2, 5),
    (wb_m, 'Auto. Verlängerungen Mai 26',    '2026-05', 3, 13, 2, 5),
    # April (Migration 040 + 041)
    (wb_f, 'Auto. Verlängerungen April 2026','2026-04', 1, 13, 2, 5),
    (wb_m, 'Auto. Verlängerungen April 26',  '2026-04', 3, 13, 2, 5),
    # Jan-März Morawitz (Migration 040)
    (wb_m, 'Auto. Verlängerungen Jan 26',    '2026-01', 3, 13, 2, 5),
    (wb_m, 'Auto. Verlängerungen Febr 26',   '2026-02', 3, 13, 2, 5),
    (wb_m, 'Auto. Verlängerungen März 26',   '2026-03', 3, 13, 2, 5),
]

conn = sqlite3.connect(DB)
lines = []
lines.append('-- Migration 042: VL auslaufend_am aus Excel-Col5 (Laufzeit-bis) wiederherstellen')
lines.append('-- + weitergeben_an_vertrieb=Ja für Verloren-VL-Deals in betroffenen Monaten')
lines.append('')

total_updated = 0
no_match = 0

for (wb, sheet_name, monat, company_id, data_start, col_kunde, col_auslaufend) in SHEETS:
    try:
        ws = wb[sheet_name]
    except KeyError:
        print(f'  WARNUNG: Sheet {sheet_name!r} nicht gefunden, übersprungen')
        continue

    sheet_rows = read_vl_auslaufend(ws, data_start, col_kunde, col_auslaufend)
    updated = 0
    not_found = 0

    lines.append(f'-- {sheet_name} → monat={monat}, company_id={company_id}')

    for r in sheet_rows:
        if not r['auslaufend_am']:
            continue
        kunde = r['kunde']
        auslaufend = r['auslaufend_am']

        # Finde passenden Deal in DB (by monat + company_id + TRIM(kunde))
        db_rows = conn.execute(
            "SELECT id FROM deals_vl WHERE monat=? AND company_id=? AND TRIM(kunde)=?",
            (monat, company_id, kunde)
        ).fetchall()

        if not db_rows:
            not_found += 1
            continue

        for db_row in db_rows:
            deal_id = db_row[0]
            lines.append(f"UPDATE deals_vl SET auslaufend_am={sql_val(auslaufend)} WHERE id={deal_id};")
            updated += 1

    print(f'  {sheet_name}: {len(sheet_rows)} Zeilen, {updated} Updates, {not_found} nicht gefunden')
    total_updated += updated
    no_match += not_found

lines.append('')
lines.append('-- weitergeben_an_vertrieb=Ja für Verloren-VL-Deals in betroffenen Monaten')
lines.append('-- (vorheriger Wert war manuell gesetzt; als Startpunkt alle auf Ja setzen)')
affected_months = "'2026-01','2026-02','2026-03','2026-04','2026-05','2026-06'"
lines.append(f"UPDATE deals_vl SET weitergeben_an_vertrieb='Ja'")
lines.append(f"  WHERE monat IN ({affected_months})")
lines.append(f"  AND status='Verloren'")
lines.append(f"  AND company_id IN (1,3)")
lines.append(f"  AND (weitergeben_an_vertrieb IS NULL OR weitergeben_an_vertrieb='');")

# Lokal anwenden
full_sql = '\n'.join(lines) + '\n'
conn.executescript('PRAGMA foreign_keys=OFF;\n' + full_sql)
conn.commit()

# Verifikation
print('\n=== Verifikation ===')
for monat in ('2026-01','2026-02','2026-03','2026-04','2026-05','2026-06'):
    r_auslaufend = conn.execute(
        "SELECT COUNT(*) FROM deals_vl WHERE monat=? AND company_id IN (1,3) AND auslaufend_am IS NOT NULL",
        (monat,)
    ).fetchone()[0]
    r_weitergeben = conn.execute(
        "SELECT COUNT(*) FROM deals_vl WHERE monat=? AND company_id IN (1,3) AND weitergeben_an_vertrieb='Ja' AND status='Verloren'",
        (monat,)
    ).fetchone()[0]
    print(f'  {monat}: {r_auslaufend} mit auslaufend_am, {r_weitergeben} Verloren+weitergeben=Ja')

conn.close()

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(full_sql)

print(f'\nGesamt: {total_updated} auslaufend_am Updates, {no_match} ohne DB-Match')
print(f'Geschrieben: {OUT}')
