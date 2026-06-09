#!/usr/bin/env python3
"""
Import deals from fach.digital Excel file into the KPI Tracker SQLite database.
- Clears all existing fach.digital (company_id=1) deals in NK, BK, VL
- Imports all months Jan-Jun 2026
- Preserves employee data unchanged
"""

import sqlite3
import openpyxl
from datetime import datetime, date
import sys
import os

EXCEL_PATH = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/Angebots-Tracker_fach.digital_2026 (3).xlsx'
DB_PATH    = '/Users/denizakpinar/Desktop/AE-Tracker_Projekt/kpi-tracker/backend/data/kpi.db'
COMPANY_ID = 1  # fach.digital

# ── Helpers ──────────────────────────────────────────────────────────────────

def clean(v):
    """Return None for empty/whitespace-only values."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return v

def status_map(v):
    """Map Excel Realisierung to app status."""
    if v is None:
        return 'Offen'
    s = str(v).strip()
    if s == 'Ja':
        return 'Gewonnen'
    if s == 'Nein':
        return 'Verloren'
    return 'Offen'  # Verhandlung, Closing Call 2, etc.

def find_header_row(sheet, keywords, max_row=25):
    """Find the row index (0-based) that contains all of the given keywords."""
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i > max_row:
            break
        row_strs = [str(v).strip().lower() if v else '' for v in row]
        if all(kw.lower() in ' '.join(row_strs) for kw in keywords):
            return i, list(row)
    return None, None

def col_idx(header, *names):
    """Return the 0-based index of the first header cell matching any of `names`."""
    for name in names:
        for i, h in enumerate(header):
            if h and name.lower() in str(h).strip().lower():
                return i
    return None

def first_name(s):
    """Extract first name from 'Albert Klaiqi, Timo Tonndorf' → 'Albert Klaiqi'."""
    if not s:
        return None
    s = str(s).strip()
    # If comma-separated, take first
    if ',' in s:
        s = s.split(',')[0].strip()
    return s if s else None

def build_emp_lookup(conn):
    """Build name→id lookups by role priority."""
    rows = conn.execute(
        "SELECT id, name, rolle, standort FROM employees WHERE aktiv=1"
    ).fetchall()

    # For each name, collect all employees with that name
    by_name = {}
    for emp_id, name, rolle, standort in rows:
        key = name.strip().lower()
        if key not in by_name:
            by_name[key] = []
        by_name[key].append({'id': emp_id, 'name': name, 'rolle': rolle, 'standort': standort})

    # Build typo/alias map (lowercase original → corrected name key)
    aliases = {
        'andeas': 'andreas',
        'serkan': None,  # not in DB, skip
        'albert klaiqi': 'albert',
        'albert/timo': 'albert',
    }

    def lookup(name, prefer_roles=None):
        if not name:
            return None
        key = name.strip().lower()

        # Check alias map
        if key in aliases:
            if aliases[key] is None:
                return None  # known unknown
            key = aliases[key]

        candidates = by_name.get(key, [])

        # Fallback: try first word only (e.g. "Albert Klaiqi" → "albert")
        if not candidates:
            first_word = key.split()[0]
            candidates = by_name.get(first_word, [])

        if not candidates:
            print(f"  ⚠️  Unknown employee: '{name}'")
            return None
        if prefer_roles:
            for r in prefer_roles:
                for c in candidates:
                    if c['rolle'] == r:
                        return c['id']
        return candidates[0]['id']

    return lookup


# ── NK Import ─────────────────────────────────────────────────────────────────

NK_SHEETS = [
    ('NK-Vertrieb Jan 26',      '2026-01'),
    ('NK-Vertrieb Februar 26',  '2026-02'),
    ('NK-Vertrieb März 26',     '2026-03'),
    ('NK-Vertrieb April 2026',  '2026-04'),
    ('NK-Vertrieb Mai BN 2026', '2026-05'),
    ('NK-Vertrieb Mai BS 2026', '2026-05'),
    ('NK-Vertrieb Juni BN 2026','2026-06'),
    ('NK-Vertrieb Juni BS 2026','2026-06'),
]

def parse_nk_sheet(sheet, monat, lookup):
    rows_out = []
    header_row_idx, header = find_header_row(sheet, ['quelle', 'datum', 'closer'], max_row=25)
    if header is None:
        print(f"  ⚠️  Could not find header row in {sheet.title}")
        return []

    # Map column names
    c_quelle     = col_idx(header, 'quelle')
    c_datum      = col_idx(header, 'datum')
    c_closer     = col_idx(header, 'closer')
    c_opener     = col_idx(header, 'opener')
    c_setter     = col_idx(header, 'setter')
    c_kunde      = col_idx(header, 'kunde')
    c_angebotsnr = col_idx(header, 'angebotsnummer')
    c_dienst     = col_idx(header, 'dienstleistung')
    c_wert       = col_idx(header, 'wert insgesamt', 'wert')
    c_laufzeit   = col_idx(header, 'laufzeit')
    c_kommentar  = col_idx(header, 'kommentar')
    c_status     = col_idx(header, 'realisierung')

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue
        if not any(v is not None for v in row):
            continue
        # Skip rows where datum is not a date/datetime
        raw_datum = row[c_datum] if c_datum is not None else None
        if raw_datum is None:
            continue
        if not isinstance(raw_datum, (datetime, date)):
            continue

        datum_str   = raw_datum.strftime('%Y-%m-%d') if isinstance(raw_datum, datetime) else str(raw_datum)
        quelle      = clean(row[c_quelle])     if c_quelle     is not None else None
        closer_name = clean(first_name(row[c_closer])) if c_closer  is not None else None
        opener_name = clean(first_name(row[c_opener])) if c_opener  is not None else None
        setter_name = clean(first_name(row[c_setter])) if c_setter  is not None else None
        kunde       = clean(row[c_kunde])      if c_kunde      is not None else None
        angebotsnr  = clean(row[c_angebotsnr]) if c_angebotsnr is not None else None
        dienst      = clean(row[c_dienst])     if c_dienst     is not None else None
        angebotswert= clean(row[c_wert])       if c_wert       is not None else None
        laufzeit    = clean(row[c_laufzeit])   if c_laufzeit   is not None else None
        kommentar   = clean(row[c_kommentar])  if c_kommentar  is not None else None
        raw_status  = row[c_status]            if c_status     is not None else None
        status      = status_map(raw_status)

        # Numeric cleanup
        try:
            angebotswert = float(angebotswert) if angebotswert is not None else None
        except (TypeError, ValueError):
            angebotswert = None
        try:
            laufzeit = float(str(laufzeit).replace('M','').strip()) if laufzeit is not None else None
        except (TypeError, ValueError):
            laufzeit = None

        # Skip if no Kunde (probably a blank or header row)
        if not kunde:
            continue

        closer_id = lookup(closer_name, prefer_roles=['NKV-Closer', 'Setter'])
        opener_id = lookup(opener_name, prefer_roles=['Opener'])
        setter_id = lookup(setter_name, prefer_roles=['Setter', 'Opener'])

        ae_wert = angebotswert if status == 'Gewonnen' else None
        gewonnen_datum = datum_str if status == 'Gewonnen' else None
        gewonnen_monat = monat    if status == 'Gewonnen' else None

        rows_out.append({
            'datum': datum_str, 'monat': monat, 'company_id': COMPANY_ID,
            'closer_id': closer_id, 'opener_id': opener_id, 'setter_id': setter_id,
            'quelle': quelle, 'kunde': kunde, 'angebotsnummer': str(angebotsnr) if angebotsnr else None,
            'dienstleistung': dienst, 'angebotswert': angebotswert,
            'laufzeit_monate': laufzeit, 'status': status, 'ae_wert': ae_wert,
            'kommentar': kommentar,
            'gewonnen_datum': gewonnen_datum, 'gewonnen_monat': gewonnen_monat,
        })

    return rows_out


# ── BK Import ─────────────────────────────────────────────────────────────────

BK_SHEETS = [
    ('BK Jan 26',    '2026-01'),
    ('BK Feb 26',    '2026-02'),
    ('BK März 26',   '2026-03'),
    ('BK April 2026','2026-04'),
    ('BK Mai 2026',  '2026-05'),
    ('BK Juni 2026', '2026-06'),
]

def parse_bk_sheet(sheet, monat, lookup):
    rows_out = []
    header_row_idx, header = find_header_row(sheet, ['datum', 'account manager', 'kunde'], max_row=20)
    if header is None:
        print(f"  ⚠️  Could not find header row in {sheet.title}")
        return []

    c_datum      = col_idx(header, 'datum')
    c_kam        = col_idx(header, 'key account manager', 'account manager')
    c_kunde      = col_idx(header, 'kunde')
    c_angebotsnr = col_idx(header, 'angebotsnummer')
    c_dienst     = col_idx(header, 'dienstleistung')
    c_wert       = col_idx(header, 'wert insgesamt', 'wert')
    c_laufzeit   = col_idx(header, 'laufzeit')
    # Kommentar: "Kommentar aktueller Stand" in older sheets, "Weitere Notiz" in newer
    c_kommentar  = col_idx(header, 'kommentar', 'weitere notiz')
    c_status     = col_idx(header, 'realisierung')

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue
        if not any(v is not None for v in row):
            continue
        raw_datum = row[c_datum] if c_datum is not None else None
        if raw_datum is None:
            continue
        if not isinstance(raw_datum, (datetime, date)):
            continue

        datum_str    = raw_datum.strftime('%Y-%m-%d') if isinstance(raw_datum, datetime) else str(raw_datum)
        kam_name     = clean(first_name(row[c_kam]))     if c_kam        is not None else None
        kunde        = clean(row[c_kunde])               if c_kunde      is not None else None
        angebotsnr   = clean(row[c_angebotsnr])          if c_angebotsnr is not None else None
        dienst       = clean(row[c_dienst])              if c_dienst     is not None else None
        angebotswert = clean(row[c_wert])                if c_wert       is not None else None
        laufzeit     = clean(row[c_laufzeit])            if c_laufzeit   is not None else None
        kommentar    = clean(row[c_kommentar])           if c_kommentar  is not None else None
        raw_status   = row[c_status]                     if c_status     is not None else None
        status       = status_map(raw_status)

        try:
            angebotswert = float(angebotswert) if angebotswert is not None else None
        except (TypeError, ValueError):
            angebotswert = None
        try:
            laufzeit = float(str(laufzeit).strip()) if laufzeit is not None else None
        except (TypeError, ValueError):
            laufzeit = None

        if not kunde:
            continue

        kam_id = lookup(kam_name, prefer_roles=['KAM', 'Multi'])
        ae_wert = angebotswert if status == 'Gewonnen' else None
        gewonnen_datum = datum_str if status == 'Gewonnen' else None
        gewonnen_monat = monat    if status == 'Gewonnen' else None

        # angebotsnummer can be a number in Excel – convert to string
        if isinstance(angebotsnr, float):
            angebotsnr = str(int(angebotsnr))
        elif angebotsnr is not None:
            angebotsnr = str(angebotsnr)

        rows_out.append({
            'datum': datum_str, 'monat': monat, 'company_id': COMPANY_ID,
            'kam_id': kam_id, 'kunde': kunde,
            'angebotsnummer': angebotsnr, 'dienstleistung': dienst,
            'angebotswert': angebotswert, 'laufzeit_monate': laufzeit,
            'status': status, 'ae_wert': ae_wert, 'kommentar': kommentar,
            'gewonnen_datum': gewonnen_datum, 'gewonnen_monat': gewonnen_monat,
        })

    return rows_out


# ── VL Import ─────────────────────────────────────────────────────────────────

VL_SHEETS = [
    ('Auto. Verlängerungen Jan 26',    '2026-01'),
    ('Auto. Verlängerungen Feb 26',    '2026-02'),
    ('Auto. Verlängerungen März 26',   '2026-03'),
    ('Auto. Verlängerungen April 2026','2026-04'),
    ('Auto. Verlängerungen Mai 2026',  '2026-05'),
    ('Auto. Verlängerungen Juni 2026', '2026-06'),
]

def safe_col(row, idx):
    """Safely get row[idx], returning None if out of range."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]

def parse_vl_sheet(sheet, monat, lookup):
    rows_out = []
    header_row_idx, header = find_header_row(sheet, ['datum', 'kunde', 'account'], max_row=20)
    if header is None:
        print(f"  ⚠️  Could not find header row in {sheet.title}")
        return []

    c_datum    = col_idx(header, 'datum')
    c_kunde    = col_idx(header, 'kunde')
    c_kam      = col_idx(header, 'account')
    c_dienst   = col_idx(header, 'leistung')
    c_laufzeit = col_idx(header, 'neue laufzeit', 'laufzeit in', 'monate')
    # AE column: try multiple names
    c_ae       = col_idx(header, 'ae durch verlängerung', 'ae durch', 'ae')
    c_status   = col_idx(header, 'realisierung')
    c_wie_vielt= col_idx(header, 'wie vielte', 'vielte')

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue
        if not any(v is not None for v in row):
            continue
        raw_datum = safe_col(row, c_datum)
        if raw_datum is None:
            continue
        if not isinstance(raw_datum, (datetime, date)):
            continue

        datum_str  = raw_datum.strftime('%Y-%m-%d') if isinstance(raw_datum, datetime) else str(raw_datum)
        kam_name   = clean(first_name(safe_col(row, c_kam)))
        kunde      = clean(safe_col(row, c_kunde))
        dienst     = clean(safe_col(row, c_dienst))
        laufzeit   = clean(safe_col(row, c_laufzeit))
        raw_ae     = safe_col(row, c_ae)
        raw_status = safe_col(row, c_status)
        wie_vielt  = safe_col(row, c_wie_vielt)
        status     = status_map(raw_status)

        try:
            ae_val = float(raw_ae) if raw_ae is not None else None
        except (TypeError, ValueError):
            ae_val = None
        try:
            laufzeit = float(str(laufzeit).strip()) if laufzeit is not None else None
        except (TypeError, ValueError):
            laufzeit = None
        try:
            wie_vielt_val = float(wie_vielt) if wie_vielt is not None and str(wie_vielt).strip() not in ('', 'X', 'x') else None
        except (TypeError, ValueError):
            wie_vielt_val = None

        if not kunde:
            continue

        kam_id = lookup(kam_name, prefer_roles=['KAM', 'Multi'])
        ae_wert = ae_val if status == 'Gewonnen' else None
        gewonnen_datum = datum_str if status == 'Gewonnen' else None
        gewonnen_monat = monat    if status == 'Gewonnen' else None

        rows_out.append({
            'datum': datum_str, 'monat': monat, 'company_id': COMPANY_ID,
            'kam_id': kam_id, 'kunde': kunde, 'dienstleistung': dienst,
            'angebotswert': ae_val,   # möglicher AE
            'ae_wert': ae_wert,       # realisierter AE (nur wenn Gewonnen)
            'laufzeit_monate': laufzeit, 'status': status,
            'wie_vielt_verlaengerung': wie_vielt_val,
            'kommentar': None,
            'gewonnen_datum': gewonnen_datum, 'gewonnen_monat': gewonnen_monat,
        })

    return rows_out


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    print("Connecting to SQLite DB...")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    lookup = build_emp_lookup(conn)

    # ── NK ────────────────────────────────────────────────────────────────────
    all_nk = []
    for sheet_name, monat in NK_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet not found: {sheet_name}")
            continue
        print(f"Parsing NK: {sheet_name} ({monat})")
        rows = parse_nk_sheet(wb[sheet_name], monat, lookup)
        print(f"  → {len(rows)} deals")
        all_nk.extend(rows)

    # ── BK ────────────────────────────────────────────────────────────────────
    all_bk = []
    for sheet_name, monat in BK_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet not found: {sheet_name}")
            continue
        print(f"Parsing BK: {sheet_name} ({monat})")
        rows = parse_bk_sheet(wb[sheet_name], monat, lookup)
        print(f"  → {len(rows)} deals")
        all_bk.extend(rows)

    # ── VL ────────────────────────────────────────────────────────────────────
    all_vl = []
    for sheet_name, monat in VL_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet not found: {sheet_name}")
            continue
        print(f"Parsing VL: {sheet_name} ({monat})")
        rows = parse_vl_sheet(wb[sheet_name], monat, lookup)
        print(f"  → {len(rows)} deals")
        all_vl.extend(rows)

    # ── Write to DB ───────────────────────────────────────────────────────────
    print(f"\nTotal to import: NK={len(all_nk)}, BK={len(all_bk)}, VL={len(all_vl)}")
    confirm = input("Delete existing fach.digital data and import? [y/N] ").strip().lower()
    if confirm != 'y':
        print("Aborted.")
        return

    cur = conn.cursor()

    print("Deleting existing fach.digital deals...")
    cur.execute("DELETE FROM deals_nk WHERE company_id = ?", (COMPANY_ID,))
    cur.execute("DELETE FROM deals_bk WHERE company_id = ?", (COMPANY_ID,))
    cur.execute("DELETE FROM deals_vl WHERE company_id = ?", (COMPANY_ID,))
    print(f"  Deleted NK={cur.rowcount} rows")

    print("Inserting NK deals...")
    for r in all_nk:
        cur.execute("""
            INSERT INTO deals_nk
              (datum,monat,company_id,closer_id,opener_id,setter_id,quelle,kunde,
               angebotsnummer,dienstleistung,angebotswert,laufzeit_monate,status,
               ae_wert,kommentar,gewonnen_datum,gewonnen_monat)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (r['datum'],r['monat'],r['company_id'],r['closer_id'],r['opener_id'],
              r['setter_id'],r['quelle'],r['kunde'],r['angebotsnummer'],
              r['dienstleistung'],r['angebotswert'],r['laufzeit_monate'],r['status'],
              r['ae_wert'],r['kommentar'],r['gewonnen_datum'],r['gewonnen_monat']))

    print("Inserting BK deals...")
    for r in all_bk:
        cur.execute("""
            INSERT INTO deals_bk
              (datum,monat,company_id,kam_id,kunde,angebotsnummer,dienstleistung,
               angebotswert,laufzeit_monate,status,ae_wert,kommentar,
               gewonnen_datum,gewonnen_monat)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (r['datum'],r['monat'],r['company_id'],r['kam_id'],r['kunde'],
              r['angebotsnummer'],r['dienstleistung'],r['angebotswert'],
              r['laufzeit_monate'],r['status'],r['ae_wert'],r['kommentar'],
              r['gewonnen_datum'],r['gewonnen_monat']))

    print("Inserting VL deals...")
    for r in all_vl:
        cur.execute("""
            INSERT INTO deals_vl
              (datum,monat,company_id,kam_id,kunde,dienstleistung,angebotswert,
               ae_wert,laufzeit_monate,status,wie_vielt_verlaengerung,kommentar,
               gewonnen_datum,gewonnen_monat)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (r['datum'],r['monat'],r['company_id'],r['kam_id'],r['kunde'],
              r['dienstleistung'],r['angebotswert'],r['ae_wert'],r['laufzeit_monate'],
              r['status'],r['wie_vielt_verlaengerung'],r['kommentar'],
              r['gewonnen_datum'],r['gewonnen_monat']))

    conn.commit()
    print("\n✅ Import complete!")

    # Summary
    nk_cnt = conn.execute("SELECT COUNT(*) FROM deals_nk WHERE company_id=?", (COMPANY_ID,)).fetchone()[0]
    bk_cnt = conn.execute("SELECT COUNT(*) FROM deals_bk WHERE company_id=?", (COMPANY_ID,)).fetchone()[0]
    vl_cnt = conn.execute("SELECT COUNT(*) FROM deals_vl WHERE company_id=?", (COMPANY_ID,)).fetchone()[0]
    print(f"   NK rows: {nk_cnt}")
    print(f"   BK rows: {bk_cnt}")
    print(f"   VL rows: {vl_cnt}")
    conn.close()

if __name__ == '__main__':
    main()
